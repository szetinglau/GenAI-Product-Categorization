import difflib
import pandas as pd
import os
import openpyxl

def save_evaluation_excel(initiated_items, approved_items, predictions, detailed_comparison, mismatch_summary, accuracy_summary, category_performance_df, hl1, hl2, hl3, output_dir):
    output_path = os.path.join(output_dir, "evaluation_results.xlsx")
    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        # Write all DataFrames to their respective sheets
        initiated_items.to_excel(writer, sheet_name="Initiated Items", index=False)
        approved_items.to_excel(writer, sheet_name="Approved Items", index=False)
        predictions.to_excel(writer, sheet_name="Predicted Items", index=False)
        detailed_comparison.to_excel(writer, sheet_name="Detailed Comparison", index=False)
        mismatch_summary.to_excel(writer, sheet_name="Mismatch Summary", index=False)
        accuracy_summary.to_excel(writer, sheet_name="Accuracy Summary", index=False)
        category_performance_df.to_excel(writer, sheet_name="Individual Hierarchy Accuracy", index=False)
        hl1.to_excel(writer, sheet_name="Hierarchy Level 1", index=False)
        hl2.to_excel(writer, sheet_name="Hierarchy Level 2", index=False)
        hl3.to_excel(writer, sheet_name="Hierarchy Level 3", index=False)

        format_eval_results(detailed_comparison, mismatch_summary, writer)

    auto_format_excel(output_path)
    print(f"Evaluation results saved to: {output_path}")

def format_eval_results(detailed_comparison, mismatch_summary, writer):
    workbook = writer.book
    detailed_ws = writer.sheets["Detailed Comparison"]
    mismatch_ws = writer.sheets["Mismatch Summary"]

    mismatch_format = workbook.add_format({'bg_color': '#FFCCCC'})
    red_format = workbook.add_format({'font_color': 'red'})
    blue_format = workbook.add_format({'font_color': 'blue'})

    for idx, row in detailed_comparison.iterrows():
        if not row["Class_Name_Match"]:
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Class_Name_predicted"), row["Class_Name_predicted"], mismatch_format)
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Class_Name_actual"), row["Class_Name_actual"], mismatch_format)
        if not row["PBH_Match"]:
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("PBH_predicted"), row["PBH_predicted"], mismatch_format)
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("PBH_actual"), row["PBH_actual"], mismatch_format)
        if not row["Analytical_Hierarchy_Match"]:
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_predicted"), row["Analytical_Hierarchy_predicted"], mismatch_format)
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_actual"), row["Analytical_Hierarchy_actual"], mismatch_format)
        if not row["Analytical_Hierarchy_CD_Match"]:
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_CD_predicted"), row["Analytical_Hierarchy_CD_predicted"], mismatch_format)
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_CD_actual"), row["Analytical_Hierarchy_CD_actual"], mismatch_format)
        if row["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"]:
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_CD_actual"), row["Analytical_Hierarchy_CD_actual"], red_format)
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_CD_Not_In_Hierarchy_File"), row["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"], mismatch_format)

    for idx, row in mismatch_summary.iterrows():
        if not row["Class_Name_Match"]:
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Class_Name_predicted"), row["Class_Name_predicted"], mismatch_format)
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Class_Name_actual"), row["Class_Name_actual"], mismatch_format)
        if not row["PBH_Match"]:
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("PBH_predicted"), row["PBH_predicted"], mismatch_format)
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("PBH_actual"), row["PBH_actual"], mismatch_format)
        if not row["Analytical_Hierarchy_Match"]:
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_predicted"), row["Analytical_Hierarchy_predicted"], mismatch_format)
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_actual"), row["Analytical_Hierarchy_actual"], mismatch_format)
        if not row["Analytical_Hierarchy_CD_Match"]:
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_CD_predicted"), row["Analytical_Hierarchy_CD_predicted"], mismatch_format)
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_CD_actual"), row["Analytical_Hierarchy_CD_actual"], mismatch_format)
        if row["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"]:
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_CD_actual"), row["Analytical_Hierarchy_CD_actual"], red_format)
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_CD_Not_In_Hierarchy_File"), row["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"], mismatch_format)

    # Conditional formatting for Mismatch Summary sheet with rich diff formatting
    for i, row in mismatch_summary.iterrows():
        for field in ["Class_Name", "PBH", "Analytical_Hierarchy", "Analytical_Hierarchy_CD"]:
            col = mismatch_summary.columns.get_loc(f"{field}_Diff")
                # For hierarchical field, apply diff highlighting if mismatch occurs
            if field == "Analytical_Hierarchy" and not row[f"{field}_Match"]:
                rich_text = highlight_hierarchy_diff(row[f"{field}_predicted"], row[f"{field}_actual"], red_format, blue_format)
                mismatch_ws.write_rich_string(i + 1, col, *rich_text)
            elif not row[f"{field}_Match"]:
                rich_text = diff_format_words(row[f"{field}_predicted"], row[f"{field}_actual"], red_format, blue_format)
                mismatch_ws.write_rich_string(i + 1, col, *rich_text)

def diff_format_words(predicted, actual, red_format, blue_format):
    # Use difflib to get a diff between actual and predicted words.
    actual_words = actual.split()
    predicted_words = predicted.split()
    diff = list(difflib.ndiff(actual_words, predicted_words))
    rich_text = []
    for token in diff:
        if token.startswith('- '):
            # Removed word (actual): highlight in red
            rich_text.extend([red_format, token[2:] + " "])
        elif token.startswith('+ '):
            # Added word (predicted): highlight in blue
            rich_text.extend([blue_format, f"{token[2:]} "])  # Removed (+) prefix and fixed color
        else:
            # Common word: normal text
            rich_text.append(token[2:] + " ")
    return rich_text

def highlight_hierarchy_diff(predicted, actual, red_format, blue_format):
    # Split each hierarchy into segments.
    pred_segs = [seg.strip() for seg in predicted.split(">>")]
    act_segs = [seg.strip() for seg in actual.split(">>")]
    rich_text = []
    # Loop over each segment and highlight differences.
    for i in range(max(len(pred_segs), len(act_segs))):
        pred_seg = pred_segs[i] if i < len(pred_segs) else ""
        act_seg = act_segs[i] if i < len(act_segs) else ""
        if pred_seg == act_seg:
            rich_text.append(act_seg + ">>")
        else:
            # Highlight diff between the segments.
            seg_diff = diff_format_words(pred_seg, act_seg, red_format, blue_format)
            rich_text.extend(seg_diff)
            if i < len(act_segs) - 1:
                rich_text.append(">>")
    # Remove trailing separator if any.
    if rich_text and rich_text[-1] == ">>":
        rich_text = rich_text[:-1]
    return rich_text

def auto_format_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_length + 2
        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "A2"
    wb.save(file_path)
    print(f"Excel file '{file_path}' formatted successfully.")
