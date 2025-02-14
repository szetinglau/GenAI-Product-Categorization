'''
# - TODO Turn off content filter for this. It triggers maybe once every 100 product items when classifying through gpt-4o: Content filter error: Could not parse response content as the request was rejected by the content filter
- TODO Need to spit out each search results + prompt going to GPT-4o-mini for verification purposes
- TODO Update Readme as this might be a highly repeatable solution
- TODO Deployable solution idea: Web UI, drop in a file in this format, columns, rows, each even row hasn't been looked at, each odd row is ground truth. Spits out the accuracy of having Azure AI Search + GPT-4o/4o-mini/o1 make the product categorization
- TODO Make a One Click Deploy, yes it'll be painful, but do it and you enhance repeatability
- TODO Test with Blanks on Class
'''

import openai
import pandas as pd
import numpy as np
import requests, json, os
import time
import logging
import openpyxl

from azure.core.exceptions import HttpResponseError
from azure.search.documents import SearchClient
from azure.search.documents.models import VectorizedQuery, VectorizableTextQuery
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import (
    SearchIndex,
    SimpleField,
    SearchField,
    SearchFieldDataType,
    VectorSearch,
    HnswAlgorithmConfiguration,
    VectorSearchProfile,
    AzureOpenAIVectorizer,
    AzureOpenAIVectorizerParameters,
    SemanticConfiguration,
    SemanticPrioritizedFields,
    SemanticField,
    SemanticSearch
)
from azure.core.credentials import AzureKeyCredential
from azure.core.pipeline.policies import RetryPolicy
from dotenv import load_dotenv
from openai import AzureOpenAI
from tenacity import retry, stop_after_attempt, wait_exponential
from pydantic import BaseModel
from datetime import datetime
import difflib

load_dotenv(override=True)

# Ensure the outputs directory exists
os.makedirs("outputs", exist_ok=True)

# # Create a timestamped folder within the outputs directory
# timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
# output_dir = os.path.join("outputs", timestamp)
# os.makedirs(output_dir, exist_ok=True)

# Use a fixed timestamp folder directory
output_dir = os.path.join("outputs", "20250205_021108")
os.makedirs(output_dir, exist_ok=True)

# Set up logging
log_file = os.path.join(output_dir, "execution.log")
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[
    logging.FileHandler(log_file),
    logging.StreamHandler()
])
logger = logging.getLogger()

# Azure OpenAI credentials
azure_openai_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")  # Replace with your Azure OpenAI endpoint
azure_openai_key = os.getenv("AZURE_OPENAI_KEY")  # Replace with your Azure OpenAI API key
azure_api_version = os.getenv("AZURE_OPENAI_API_VERSION")

openai_deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME")
openai_model_name = os.getenv("AZURE_OPENAI_MODEL_NAME")

text_embedding_deployment_name = os.getenv("AZURE_OPENAI_TEXT_EMBEDDING_DEPLOYMENT_NAME")
text_embedding_model_name = os.getenv("AZURE_OPENAI_TEXT_EMBEDDING_MODEL_NAME")

# Initialize the Azure OpenAI client
aoai_client = AzureOpenAI(
            azure_endpoint=azure_openai_endpoint,
            api_version=azure_api_version,
            api_key=azure_openai_key,
        )

# Azure Cognitive Search credentials
search_service_endpoint = os.getenv("SEARCH_SERVICE_ENDPOINT")  # Replace with your Azure Cognitive Search endpoint
search_api_key = os.getenv("SEARCH_API_KEY")  # Replace with your Azure Cognitive Search API key
index_name = os.getenv("SEARCH_INDEX_NAME")  # Replace with your Azure Cognitive Search index name
search_client = SearchClient(endpoint=search_service_endpoint, index_name=index_name, credential=AzureKeyCredential(search_api_key))

# Define the fields to be used for embedding and search
fields = [
    "Item_Num", "Description_1", "Description_2", "GTIN", "Brand_Id", "Brand_Name", 
    "GDSN_Brand", "Long_Product_Name", "Pack", "Size", "Size_UOM", "Class_Id", 
    "Class_Name", "PBH_ID", "PBH", "Analytical_Hierarchy_CD", "Analytical_Hierarchy", 
    "Temp_Min", "Temp_Max", "Benefits", "General_Description"
]

# Load the input data
def load_data(class_pbh_file, item_hierarchy_file, product_sku_items_file):
    # Load data into pandas DataFrames
    class_pbh_df = pd.read_excel(class_pbh_file, engine="openpyxl")
    item_hierarchy_df = pd.read_excel(item_hierarchy_file, engine="openpyxl")
    product_sku_items_df = pd.read_excel(product_sku_items_file, engine="openpyxl")

    # Log the size of each DataFrame
    logger.info(f"Class PBH DataFrame size: {class_pbh_df.shape}")
    logger.info(f"Item Hierarchy DataFrame size: {item_hierarchy_df.shape}")
    logger.info(f"Product SKU Items DataFrame size: {product_sku_items_df.shape}")

    # Log unique values in the Status column
    logger.info(f"Unique values in Status column: {product_sku_items_df['Status'].unique()}")

    # Log the length of the item hierarchy DataFrame before filtering
    logger.info(f"Item Hierarchy DataFrame length before filtering: {len(item_hierarchy_df)}")

    # Filter rows based on column C
    item_hierarchy_df = item_hierarchy_df[item_hierarchy_df.iloc[:, 2] == 'USE THESE VALUES']

    # Log the length of the item hierarchy DataFrame after filtering
    logger.info(f"Item Hierarchy DataFrame length after filtering: {len(item_hierarchy_df)}")

    # Convert to string and remove leading/trailing whitespaces prior to merging
    # product_sku_items_df["Hierarchy CD"] = product_sku_items_df["Hierarchy CD"].astype(str).str.strip()
    # item_hierarchy_df["Hierarchy CD"] = item_hierarchy_df["Hierarchy CD"].astype(str).str.strip()

    logger.info("Processing approved items...")
    approved_items_df = filter_and_merge_items(product_sku_items_df, 'APPROVED', item_hierarchy_df, output_dir)
    
    logger.info("Processing initiated items...")
    initiated_items_df = filter_and_merge_items(product_sku_items_df, 'INITIATED', item_hierarchy_df, output_dir)
    
    # Clear specified classification fields in initiated_items_df
    for col in ["Class_Name", "PBH", "Analytical_Hierarchy", "Analytical_Hierarchy_CD"]:
        if col in initiated_items_df.columns:
            initiated_items_df[col] = ""
            
    # Check if there are any non-empty rows in the specified columns of initiated_items_df
    non_empty_rows = initiated_items_df[
        (initiated_items_df["Class_Name"] != "") |
        (initiated_items_df["PBH"] != "") |
        (initiated_items_df["Analytical_Hierarchy"] != "") |
        (initiated_items_df["Analytical_Hierarchy_CD"] != "")
    ]

    # Log a warning if there are any non-empty rows
    if not non_empty_rows.empty:
        logger.warning(f"Found {len(non_empty_rows)} rows in initiated items with non-empty classification fields.")
    else:
        logger.info("All specified classification fields in initiated items are empty.")
    
    logger.info(f"Number of approved items: {len(approved_items_df)}")
    logger.info(f"Number of initiated items: {len(initiated_items_df)}")

    logger.info("Data loaded successfully.")
    return class_pbh_df, item_hierarchy_df, approved_items_df, initiated_items_df

# Filter items based on status
def filter_items_by_status(items, status):
    return items[items['Status'] == status][[
        'Item_Num', 'Description_1', 'Description_2', 'GTIN', 'Brand_Id', 'Brand_Name', 
        'GDSN_Brand', 'Long_Product_Name', 'Pack', 'Size', 'Size_UOM', 'Class_Id', 
        'Class_Name', 'PBH_ID', 'PBH', 'Hierarchy CD', 'Temp_Min', 'Temp_Max', 
        'Benefits', 'General_Description'
    ]]


def merge_items_with_hierarchy(filtered_items, item_hierarchy, output_dir, status):
    merged_data = pd.merge(
        filtered_items,
        item_hierarchy,
        left_on="Hierarchy CD",
        right_on="Hierarchy CD",
        # on="Hierarchy CD",
        how="left"
    )
    
    # Only take items from the product sku items that have hierarchies that are in the item_hierarchy reference file
    missing = merged_data[merged_data['Hierarchy Detail'].isnull()]
    if not missing.empty:
        logger.warning("Warning: The following items have missing hierarchies in the reference file:")
        logger.warning(missing[['Item_Num', 'Hierarchy CD']])
        missing_csv_path = os.path.join(output_dir, f"{status.lower()}_items_missing_hierarchies.csv")
        missing[['Item_Num', 'Hierarchy CD']].to_csv(missing_csv_path, index=False)
    return merged_data


# Rename and reorder columns
def rename_and_reorder_columns(merged_data):
    merged_data.rename(columns={
        'Hierarchy CD': 'Analytical_Hierarchy_CD',
        'Hierarchy Detail': 'Analytical_Hierarchy'
    }, inplace=True)
    merged_data = merged_data[fields]
    return merged_data


# Filter and merge items based on status and hierarchy
def filter_and_merge_items(items, status, item_hierarchy, output_dir):
    filtered_items = filter_items_by_status(items, status)
    merged_data = merge_items_with_hierarchy(filtered_items, item_hierarchy, output_dir, status)
    final_data = rename_and_reorder_columns(merged_data)
    return final_data


# Generate embeddings for each item
@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=5, max=60))
def generate_embedding(text):
    """
    Generate embeddings using Azure OpenAI with retry logic
    """
    response = aoai_client.embeddings.create(input=[text], model=text_embedding_model_name)
    return response.data[0].embedding


def save_documents_to_json(documents, filename):
    with open(os.path.join(output_dir, filename), 'w') as f:
        json.dump(documents, f)
    print(f"Documents saved to {filename}")


def load_documents_from_json(filename):
    file_path = os.path.join(output_dir, filename)  # Use output_dir
    with open(file_path, 'r') as f:
        documents = json.load(f)
    print(f"Documents loaded from {file_path}")
    return documents


def clean_value(value):
    if pd.isna(value) or value is np.nan:
        return ""
    if isinstance(value, str):
        # Replace single quotes with double quotes and strip trailing whitespace
        return value.strip().replace("'", '"')
    return str(value)


def create_embeddings_and_documents(product_data):
    print(f"Preparing {len(product_data)} documents for upload...")
    documents = []
    # Filter the fields to be used for embedding
    embedding_fields = [
        "Item_Num", "Brand_Id", "Brand_Name", "GDSN_Brand", "Long_Product_Name", 
        "Class_Id", "Class_Name", "PBH_ID", "PBH", "Analytical_Hierarchy_CD", 
        "Analytical_Hierarchy", "Temp_Min", "Temp_Max", "Benefits", "General_Description"
    ]

    for idx, row in product_data.iterrows():
        # if idx >= 100:
        #     break

        text_to_embed = " ".join(clean_value(row[field]) for field in embedding_fields)

        document = {"id": clean_value(row["Item_Num"])}
        document.update({field: clean_value(row[field]) for field in fields})
        document["embedding"] = generate_embedding(text_to_embed)
        
        documents.append(document)
        
        if idx % 100 == 0:
            print(f"Processed {idx} documents")
            
    print(f"Finished preparing {len(documents)} documents for upload.")
    
    # Save documents to JSON
    save_documents_to_json(documents, 'prepared_documents.json')
    
    return documents


def create_semantic_search_config():
    return SemanticSearch(
        configurations=[SemanticConfiguration(
            name="mySemanticConfig",
            prioritized_fields=SemanticPrioritizedFields(
                title_field=SemanticField(field_name="Long_Product_Name"),
                content_fields=[
                    SemanticField(field_name="Brand_Name"),
                    SemanticField(field_name="GDSN_Brand"),
                    SemanticField(field_name="Long_Product_Name"),
                    SemanticField(field_name="Class_Name"),
                    SemanticField(field_name="PBH"),
                    SemanticField(field_name="Analytical_Hierarchy"),
                    SemanticField(field_name="Benefits"),
                    SemanticField(field_name="General_Description")
                ],
                keywords_fields=[
                    SemanticField(field_name="Long_Product_Name"),
                    SemanticField(field_name="Class_Name"),
                    SemanticField(field_name="PBH"),
                    SemanticField(field_name="Analytical_Hierarchy")
                ]
            )
        )]
    ) 


def create_vector_search_config():
    return VectorSearch(
        algorithms=[
            HnswAlgorithmConfiguration(
                name="myHnsw",
                # efConstruction=200,  # Larger value improves recall but increases indexing time
                # m=64                # Number of neighbors, higher values improve accuracy at the cost of memory
            )
        ],
        profiles=[
            VectorSearchProfile(
                name="myHnswProfile",
                algorithm_configuration_name="myHnsw",
                vectorizer_name="myVectorizer"
            )
        ],
        vectorizers=[
            AzureOpenAIVectorizer(
                vectorizer_name="myVectorizer",
                parameters=AzureOpenAIVectorizerParameters(
                    resource_url=azure_openai_endpoint,
                    deployment_name=text_embedding_deployment_name,
                    model_name=text_embedding_model_name,
                    api_key=azure_openai_key
                )
            )
        ]
    )

def upload_documents_to_search(documents, search_client):
    batch_size = 15
    total_batches = (len(documents) + batch_size - 1) // batch_size  # Calculate total number of batches
    successful_uploads = 0

    for i in range(0, len(documents), batch_size):
        batch = documents[i:i + batch_size]
        batch_ids = ", ".join([str(doc["id"]) for doc in batch])
        logger.info(f"Uploading batch {i // batch_size + 1}/{total_batches} with document IDs: {batch_ids}")
        try:
            # Upload the batch
            response = search_client.upload_documents(documents=batch)
            uploaded_ids = ", ".join([str(doc["id"]) for doc in batch])
            successful_uploads += len(batch)
            logger.info(f"Uploaded batch {i // batch_size + 1}/{total_batches} successfully. Batch size: {len(batch)}. Document IDs: {uploaded_ids}")
        except HttpResponseError as e:
            logger.error(f"Error uploading batch {i // batch_size + 1}/{total_batches}: {e}")
            logger.error(f"Problematic batch document IDs: {batch_ids}")
            continue

    logger.info(f"Embedding index created and documents uploaded successfully. Total successful uploads: {successful_uploads}/{len(documents)}")


# Create an index in Azure Cognitive Search
def create_embedding_index(product_data, search_client):

    """
    # TODO Enrich the item hierarchy data:
    To improve the scoring for similar Analytical Hierarchies, we can add more detailed context in column C. This context should contain specific product descriptions or characteristics that distinguish each hierarchy. Example based on the current data:
        Example for Column C (Detailed Context)
        
            Hierarchy CD    Hierarchy Detail                                    Detailed Context (Column C)
            315             BAKERY, FROZEN>>PIES/TARTS>>TARTS                   Pre-baked or ready-to-bake tarts; commonly used in dessert preparation.
            2341            BAKERY, FROZEN>>PIZZA CRUSTS>>DOUGH BALL            Unshaped pizza dough; requires manual rolling and shaping for pizza preparation.
            28              BAKERY, FROZEN>>PIZZA CRUSTS>>PAR-BAKED             Partially baked crusts; ready for toppings and quick oven finishing.
            845             BAKERY, FROZEN>>SANDWICH CARRIERS>>HOT DOG/BRAT BUN Buns tailored for hot dogs and bratwurst; includes standard and jumbo sizes.
            1307            BAKERY, FROZEN>>SANDWICH CARRIERS>>KAISER & DELI>>PAR-BAKED
                            Par-baked Kaiser rolls; used in deli sandwiches, ready for quick baking.
        
        Explanation
        The added context in column C:
            • Distinguishes similar hierarchies by describing their specific use cases or unique characteristics.
            • Provides GPT-4o and Azure AI Search with richer information to refine scoring and improve categorization.
        Ensures better alignment between ambiguous product descriptions and the correct hierarchy during categorization.
    
    """

    # Prepare the documents for upload
    try:
        # Try to load documents from JSON if they exist
        documents = load_documents_from_json('prepared_documents.json')
    except FileNotFoundError:
        # If JSON file does not exist, prepare documents and save them
        documents = create_embeddings_and_documents(product_data)

    # Define the semantic profile configuration
    semantic_search = create_semantic_search_config()

    # Configure the vector search configuration  
    vector_search = create_vector_search_config()

    # Define the embedding index schema
    index_schema = SearchIndex(
        name=index_name,
        fields=[
            # Define the key field
            SimpleField(name="id", type=SearchFieldDataType.String, key=True, sortable=True, filterable=True, facetable=True),
            
            # Fields for search and filtering
            SimpleField(name="Item_Num", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SearchField(name="Description_1", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True, analyzer_name="keyword"),
            SearchField(name="Description_2", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True, analyzer_name="keyword"),
            SimpleField(name="GTIN", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SimpleField(name="Brand_Id", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SearchField(name="Brand_Name", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True, analyzer_name="keyword"),
            SearchField(name="GDSN_Brand", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True, analyzer_name="keyword"),
            SearchField(name="Long_Product_Name", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True, analyzer_name="keyword"),
            SimpleField(name="Pack", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SimpleField(name="Size", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SimpleField(name="Size_UOM", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SimpleField(name="Class_Id", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SearchField(name="Class_Name", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True, analyzer_name="keyword"),
            SimpleField(name="PBH_ID", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SearchField(name="PBH", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True, analyzer_name="keyword"),
            SimpleField(name="Analytical_Hierarchy_CD", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SearchField(name="Analytical_Hierarchy", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True, analyzer_name="keyword"),
            SimpleField(name="Temp_Min", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SimpleField(name="Temp_Max", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True),
            SearchField(name="Benefits", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True, analyzer_name="keyword"),
            SearchField(name="General_Description", type=SearchFieldDataType.String, sortable=True, filterable=True, facetable=True, analyzer_name="keyword"),
            
            # Vector embedding field
            SearchField(name="embedding", type=SearchFieldDataType.Collection(SearchFieldDataType.Single), vector_search_dimensions=3072, vector_search_profile_name="myHnswProfile")
        ],
        vector_search=vector_search,
        semantic_search=semantic_search
    )

    # Create the index in Azure Cognitive Search
    search_index_client = SearchIndexClient(search_client._endpoint, search_client._credential)
    search_index_client.create_or_update_index(index=index_schema)
    print(f"Index '{index_name}' created successfully.")

    upload_documents_to_search(documents, search_client)

def load_item_hierarchy(file_path="Item Hierarchy 11.14.2024.xlsx"):
    """
    Load and normalize the item hierarchy DataFrame.
    """
    hierarchy = pd.read_excel(file_path, engine="openpyxl")
    hierarchy["Hierarchy CD"] = hierarchy["Hierarchy CD"].apply(
        lambda x: str(int(x)) if pd.notnull(x) and isinstance(x, float) and x.is_integer() else str(x).strip()
    )
    return hierarchy


def merge_with_item_hierarchy(predicted_items, item_hierarchy):
    """
    Merge predicted items with the item hierarchy to add the Analytical_Hierarchy_CD.
    """
    merged = pd.merge(
        predicted_items,
        item_hierarchy[['Hierarchy CD', 'Hierarchy Detail']],
        left_on='Analytical_Hierarchy',
        right_on='Hierarchy Detail',
        how='left'
    ).rename(columns={'Hierarchy CD': 'Analytical_Hierarchy_CD'}).drop('Hierarchy Detail', axis=1)

    merged["Analytical_Hierarchy_CD"] = merged["Analytical_Hierarchy_CD"].apply(
        lambda x: str(int(x)) if pd.notnull(x) and isinstance(x, float) and x.is_integer() else str(x).strip()
    )
    logger.info(f"Predictions shape after merge: {merged.shape}")
    logger.info(f"Null Analytical_Hierarchy_CD count: {merged['Analytical_Hierarchy_CD'].isnull().sum()}")
    return merged


def prepare_detailed_comparison(initiated_items, predictions, approved_items):
    """
    Create a detailed comparison DataFrame comparing predicted and actual values.
    """
    # Base details from initiated items
    detailed = initiated_items[["Item_Num", "Brand_Name", "Long_Product_Name"]].copy()

    for field in ["Class_Name", "PBH", "Analytical_Hierarchy", "Analytical_Hierarchy_CD"]:
        detailed[f"{field}_predicted"] = predictions[field].fillna("").astype(str).str.strip()
        detailed[f"{field}_actual"] = approved_items[field].fillna("").astype(str).str.strip()
        detailed[f"{field}_Match"] = detailed[f"{field}_predicted"] == detailed[f"{field}_actual"]

    # Mark Analytical_Hierarchy_CD not in hierarchy file (will be updated later)
    detailed["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"] = False

    return detailed


def create_mismatch_summary(detailed_comparison, item_hierarchy):
    """
    Create a mismatch summary DataFrame and add diff columns.
    """
    # Update Analytical_Hierarchy_CD_Not_In_Hierarchy_File based on hierarchy values
    detailed_comparison["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"] = detailed_comparison.apply(
        lambda row: row["Analytical_Hierarchy_CD_actual"] not in item_hierarchy["Hierarchy CD"].values, axis=1
    )

    # Filter to rows with any mismatch
    mismatch = detailed_comparison.loc[
        ~detailed_comparison[[
            "Class_Name_Match", "PBH_Match", "Analytical_Hierarchy_Match", "Analytical_Hierarchy_CD_Match"
        ]].all(axis=1)
    ].reset_index(drop=True)

    mismatch["Mismatch_Type"] = mismatch.apply(
        lambda row: ", ".join(
            filter(None, [
                "Class_Name" if not row["Class_Name_Match"] else "",
                "PBH" if not row["PBH_Match"] else "",
                "Analytical_Hierarchy" if not row["Analytical_Hierarchy_Match"] else "",
                "Analytical_Hierarchy_CD" if not row["Analytical_Hierarchy_CD_Match"] else "",
                "Missing Analytical_Hierarchy_CD" if row["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"] else ""
            ])
        ),
        axis=1
    )

    # Add diff columns for each field
    for field in ["Class_Name", "PBH", "Analytical_Hierarchy", "Analytical_Hierarchy_CD"]:
        mismatch[f"{field}_Diff"] = mismatch.apply(
            lambda row: diff_format_rich(row[f"{field}_predicted"], row[f"{field}_actual"]), axis=1
        )

    # Reorder columns: base fields then predicted, diff, actual columns
    base_cols = ["Item_Num", "Brand_Name", "Long_Product_Name"]
    field_order = []
    for field in ["Class_Name", "PBH", "Analytical_Hierarchy", "Analytical_Hierarchy_CD"]:
        field_order.extend([f"{field}_predicted", f"{field}_Diff", f"{field}_actual"])
    extra_cols = [col for col in mismatch.columns if col not in base_cols + field_order]
    mismatch = mismatch[base_cols + field_order + extra_cols]

    return mismatch


def compute_accuracies(detailed_comparison_df):
    # ...existing accuracy summary calculations...
    class_accuracy = (detailed_comparison_df["Class_Name_predicted"] == detailed_comparison_df["Class_Name_actual"]).mean()
    pbh_accuracy = (detailed_comparison_df["PBH_predicted"] == detailed_comparison_df["PBH_actual"]).mean()
    analytical_accuracy = (detailed_comparison_df["Analytical_Hierarchy_predicted"] == detailed_comparison_df["Analytical_Hierarchy_actual"]).mean()

    accuracy_summary = pd.DataFrame({
        "Metric": ["Class Accuracy", "PBH Accuracy", "Analytical Hierarchy Accuracy", "Analytical Hierarchy Not In File Percentage"],
        "Value": [
            class_accuracy,
            pbh_accuracy,
            analytical_accuracy,
            detailed_comparison_df["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"].mean()
        ]
    })
    
    # Build Category Performance Summary
    category_counts = detailed_comparison_df.groupby("Analytical_Hierarchy_actual")["Item_Num"].count()
    correct_classifications = detailed_comparison_df[detailed_comparison_df["Analytical_Hierarchy_Match"] == True].groupby("Analytical_Hierarchy_actual")["Item_Num"].count()
    category_performance_df = pd.DataFrame({
        "Total Items": category_counts,
        "Correctly Classified": correct_classifications
    }).fillna(0)
    category_performance_df["Accuracy (%)"] = (category_performance_df["Correctly Classified"] / category_performance_df["Total Items"]) * 100
    category_performance_df = category_performance_df.reset_index().rename(columns={"Analytical_Hierarchy_actual": "Analytical_Hierarchy"})
    category_performance_df = category_performance_df.sort_values(by="Accuracy (%)", ascending=False)

    # Compute hierarchy level accuracies
    hl1 = compute_hierarchy_level_accuracy(detailed_comparison_df, 1)
    hl2 = compute_hierarchy_level_accuracy(detailed_comparison_df, 2)
    hl3 = compute_hierarchy_level_accuracy(detailed_comparison_df, 3)

    return class_accuracy, pbh_accuracy, analytical_accuracy, accuracy_summary, category_performance_df, hl1, hl2, hl3

def compute_hierarchy_level_accuracy(detailed_comparison_df, level):
    """
    Compute breakdown at the specified hierarchy level.
    """
    column_name = f"Hierarchy_Level_{level}"
    detailed_comparison_df[column_name] = detailed_comparison_df["Analytical_Hierarchy_actual"].apply(
        lambda x: ">>".join(str(x).split(">>")[:level]) if pd.notnull(x) and len(str(x).split(">>")) >= level else "Unknown"
    )
    hl = detailed_comparison_df.groupby(column_name).agg(
        Total_Items=("Item_Num", "count"),
        Correctly_Classified=("Analytical_Hierarchy_Match", "sum")
    )
    hl = hl.reset_index()  # reset the index so the column becomes a proper column
    hl["Accuracy (%)"] = (hl["Correctly_Classified"] / hl["Total_Items"]) * 100
    hl = hl.sort_values(by="Accuracy (%)", ascending=False)
    hl = hl[[column_name, "Total_Items", "Correctly_Classified", "Accuracy (%)"]]
    return hl


def save_evaluation_excel(initiated_items, approved_items, predictions, detailed_comparison, mismatch_summary, accuracy_summary, category_performance_df, hl1, hl2, hl3):
    """
    Write the evaluation results to an Excel file with multiple tabs and apply conditional formatting.
    """
    output_path = os.path.join(output_dir, "evaluation_results.xlsx")
    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        initiated_items.to_excel(writer, sheet_name="Initiated Items", index=False)
        approved_items.to_excel(writer, sheet_name="Approved Items", index=False)
        predictions.to_excel(writer, sheet_name="Predicted Items", index=False)
        detailed_comparison.to_excel(writer, sheet_name="Detailed Comparison", index=False)
        mismatch_summary.to_excel(writer, sheet_name="Mismatch Summary", index=False)
        accuracy_summary.to_excel(writer, sheet_name="Accuracy Summary", index=False)
        category_performance_df.to_excel(writer, sheet_name="Individual Hierarchy Accuracy", index=False)
        hl1.to_excel(writer, sheet_name="Hierarchy Level 1", index=False)
        hl2.to_excel(writer, sheet_name="Hierarchy Level 2", index=False)
        hl3.to_excel(writer, sheet_name="Hierarchy Level 3", index=False)

        format_eval_results(detailed_comparison, mismatch_summary, writer)

    # After writing, apply auto-formatting to the Excel file
    auto_format_excel(output_path)
    print(f"Evaluation results saved to: {output_path}")


def format_eval_results(detailed_comparison, mismatch_summary, writer):
    workbook = writer.book
    detailed_ws = writer.sheets["Detailed Comparison"]
    mismatch_ws = writer.sheets["Mismatch Summary"]

        # Define formats
    mismatch_format = workbook.add_format({'bg_color': '#FFCCCC'})
    red_format = workbook.add_format({'font_color': 'red'})
    blue_format = workbook.add_format({'font_color': 'blue'})

        # Apply the format to mismatched cells in Detailed Comparison
    for idx, row in detailed_comparison.iterrows():
        if not row["Class_Name_Match"]:
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Class_Name_predicted"), row["Class_Name_predicted"], mismatch_format)
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Class_Name_actual"), row["Class_Name_actual"], mismatch_format)
        if not row["PBH_Match"]:
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("PBH_predicted"), row["PBH_predicted"], mismatch_format)
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("PBH_actual"), row["PBH_actual"], mismatch_format)
        if not row["Analytical_Hierarchy_Match"]:
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_predicted"), row["Analytical_Hierarchy_predicted"], mismatch_format)
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_actual"), row["Analytical_Hierarchy_actual"], mismatch_format)
        if not row["Analytical_Hierarchy_CD_Match"]:
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_CD_predicted"), row["Analytical_Hierarchy_CD_predicted"], mismatch_format)
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_CD_actual"), row["Analytical_Hierarchy_CD_actual"], mismatch_format)
        if row["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"]:
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_CD_actual"), row["Analytical_Hierarchy_CD_actual"], red_format)
            detailed_ws.write(idx + 1, detailed_comparison.columns.get_loc("Analytical_Hierarchy_CD_Not_In_Hierarchy_File"), row["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"], mismatch_format)

        # Apply the format to mismatched cells in Mismatch Summary
    for idx, row in mismatch_summary.iterrows():
        if not row["Class_Name_Match"]:
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Class_Name_predicted"), row["Class_Name_predicted"], mismatch_format)
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Class_Name_actual"), row["Class_Name_actual"], mismatch_format)
        if not row["PBH_Match"]:
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("PBH_predicted"), row["PBH_predicted"], mismatch_format)
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("PBH_actual"), row["PBH_actual"], mismatch_format)
        if not row["Analytical_Hierarchy_Match"]:
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_predicted"), row["Analytical_Hierarchy_predicted"], mismatch_format)
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_actual"), row["Analytical_Hierarchy_actual"], mismatch_format)
        if not row["Analytical_Hierarchy_CD_Match"]:
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_CD_predicted"), row["Analytical_Hierarchy_CD_predicted"], mismatch_format)
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_CD_actual"), row["Analytical_Hierarchy_CD_actual"], mismatch_format)
        if row["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"]:
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_CD_actual"), row["Analytical_Hierarchy_CD_actual"], red_format)
            mismatch_ws.write(idx + 1, mismatch_summary.columns.get_loc("Analytical_Hierarchy_CD_Not_In_Hierarchy_File"), row["Analytical_Hierarchy_CD_Not_In_Hierarchy_File"], mismatch_format)

    # Conditional formatting for Mismatch Summary sheet with rich diff formatting
    for i, row in mismatch_summary.iterrows():
        for field in ["Class_Name", "PBH", "Analytical_Hierarchy", "Analytical_Hierarchy_CD"]:
            col = mismatch_summary.columns.get_loc(f"{field}_Diff")
                # For hierarchical field, apply diff highlighting if mismatch occurs
            if field == "Analytical_Hierarchy" and not row[f"{field}_Match"]:
                rich_text = highlight_hierarchy_diff(row[f"{field}_predicted"], row[f"{field}_actual"], red_format, blue_format)
                mismatch_ws.write_rich_string(i + 1, col, *rich_text)
            elif not row[f"{field}_Match"]:
                rich_text = diff_format_words(row[f"{field}_predicted"], row[f"{field}_actual"], red_format, blue_format)
                mismatch_ws.write_rich_string(i + 1, col, *rich_text)


def evaluate_predictions(initiated_items, predicted_items, approved_items):
    """
    Evaluate predictions against ground truth and save evaluation metrics and detailed comparisons to an Excel file.
    Tabs:
        - Initiated Items
        - Approved Items
        - Predicted Items
        - Detailed Comparison
        - Mismatch Summary
        - Accuracy Summary
    """
    # Load and normalize item hierarchy
    item_hierarchy_df = load_item_hierarchy()

    # Merge predictions with hierarchy to map Analytical_Hierarchy_CD
    predictions_df = merge_with_item_hierarchy(predicted_items, item_hierarchy_df)

    # Prepare detailed comparison DataFrame
    detailed_comparison_df = prepare_detailed_comparison(initiated_items, predictions_df, approved_items)

    # Create mismatch summary DataFrame
    mismatch_summary_df = create_mismatch_summary(detailed_comparison_df, item_hierarchy_df)

    # Replace NaN and Inf values
    detailed_comparison_df = detailed_comparison_df.replace([np.nan, np.inf, -np.inf], '')
    mismatch_summary_df = mismatch_summary_df.replace([np.nan, np.inf, -np.inf], '')
    
    # Compute accuracies and format summary
    class_acc, pbh_acc, analytical_acc, accuracy_summary, category_performance_df, hl1, hl2, hl3 = compute_accuracies(detailed_comparison_df)

    # Write all results to an Excel file with multiple tabs and conditional formatting
    save_evaluation_excel(initiated_items, approved_items, predictions_df, detailed_comparison_df,
                            mismatch_summary_df, accuracy_summary, category_performance_df, hl1, hl2, hl3)

    return {
        "Class Accuracy": class_acc,
        "PBH Accuracy": pbh_acc,
        "Analytical Hierarchy Accuracy": analytical_acc
    }
    

def diff_format_rich(predicted, actual):
    diff = difflib.ndiff(actual.split(), predicted.split())
    rich_text = ""
    for token in diff:
        if token.startswith('- '):
            rich_text += f"<span style='color:red'>{token[2:]}</span> "
        elif token.startswith('+ '):
            rich_text += f"<span style='color:blue'>{token[2:]}</span> "
        else:
            rich_text += token[2:] + " "
    return rich_text.strip()


def diff_format_words(predicted, actual, red_format, blue_format):
    # Use difflib to get a diff between actual and predicted words.
    actual_words = actual.split()
    predicted_words = predicted.split()
    diff = list(difflib.ndiff(actual_words, predicted_words))
    rich_text = []
    for token in diff:
        if token.startswith('- '):
            # Removed word: highlight it in red.
            rich_text.extend([blue_format, token[2:] + " "])
        elif token.startswith('+ '):
            # Added word: can be ignored or indicated as additional.
            rich_text.extend([red_format, f"(+{token[2:]}) "])
        else:
            # Common word: normal text.
            rich_text.append(token[2:] + " ")
    return rich_text

def highlight_hierarchy_diff(predicted, actual, red_format, blue_format):
    # Split each hierarchy into segments.
    pred_segs = [seg.strip() for seg in predicted.split(">>")]
    act_segs = [seg.strip() for seg in actual.split(">>")]
    rich_text = []
    # Loop over each segment and highlight differences.
    for i in range(max(len(pred_segs), len(act_segs))):
        pred_seg = pred_segs[i] if i < len(pred_segs) else ""
        act_seg = act_segs[i] if i < len(act_segs) else ""
        if pred_seg == act_seg:
            rich_text.append(act_seg + ">>")
        else:
            # Highlight diff between the segments.
            seg_diff = diff_format_words(pred_seg, act_seg, red_format, blue_format)
            rich_text.extend(seg_diff)
            if i < len(act_segs) - 1:
                rich_text.append(">>")
    # Remove trailing separator if any.
    if rich_text and rich_text[-1] == ">>":
        rich_text = rich_text[:-1]
    return rich_text

def auto_format_excel(file_path):
    # Automatically adjust column width, enable filters and freeze the top row
    wb = openpyxl.load_workbook(file_path)
    for sheet in wb.worksheets:
        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_length + 2
        # Enable AutoFilter for the range
        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions
        # Freeze the top row
        sheet.freeze_panes = "A2"
    wb.save(file_path)
    print(f"Excel file '{file_path}' formatted successfully.")

def construct_query_text(row):
    """
    Construct a query text based on the semantic configuration.
    """
    # Combine prioritized fields for the query
    query_parts = [
        str(row.get('Long_Product_Name', '')),  # Title field
        str(row.get('Class_Name', '')),         # Content field
        str(row.get('PBH', '')),                # Content field
        str(row.get('Analytical_Hierarchy', '')), # Keywords field
        str(row.get('Brand_Name', '')),         # Content field
        str(row.get('GDSN_Brand', '')),         # Content field
        str(row.get('Benefits', '')),           # Content field
        str(row.get('General_Description', '')) # Content field
    ]
    # Remove empty parts and join with a space, excluding 'nan' values
    query_text = " ".join(part for part in query_parts if part.strip() and part.lower() != 'nan')
    return query_text


def hybrid_search(row, search_client, top_k=5):
    """
    Perform a hybrid search for the given row.
    """
    # Construct query text
    query_text = construct_query_text(row)
    
    # select=["Class_Name", "PBH", "Analytical_Hierarchy", "Analytical_Hierarchy_CD"],
    
    fields = [
        "Item_Num", "Description_1", "Description_2", "GTIN", "Brand_Id", "Brand_Name", 
        "GDSN_Brand", "Long_Product_Name", "Pack", "Size", "Size_UOM", "Class_Id", 
        "Class_Name", "PBH_ID", "PBH", "Analytical_Hierarchy_CD", "Analytical_Hierarchy", 
        "Temp_Min", "Temp_Max", "Benefits", "General_Description"
    ]

    # Define vector query
    vector_query = VectorizableTextQuery(
        text=query_text,
        k_nearest_neighbors=50,
        fields="embedding"
    )
    
    # Perform hybrid search
    results = search_client.search(
        search_text=query_text,
        semantic_configuration_name="mySemanticConfig",
        vector_queries=[vector_query],
        select=fields,
        top=top_k
    )

    # # Log results for debugging
    # for result in results:
    #     print(f"Class_Name: {result['Class_Name']}, PBH: {result['PBH']}, Analytical_Hierarchy: {result['Analytical_Hierarchy']}, Score: {result['@search.score']}")
    
    return results


class ProductClassification(BaseModel):
    class_name: str
    pbh: str
    analytical_hierarchy: str


def build_classification_prompt(row, search_results):
    """
    Build the GPT-4 prompt based on the product details and search results.
    """
    # Clear classification fields before prompting
    row["Class_Name"] = ""
    row["PBH"] = ""
    row["Analytical_Hierarchy"] = ""
    
    product_details = "### Product Details ###\n"
    for field in fields:
        product_details += f"- {field.replace('_', ' ').title()}: {row.get(field, '')}\n"
    
    prompt = f"""
You are a **product classification expert** tasked with assigning the most accurate **Class Name, PBH (Product Book Handler), and Analytical Hierarchy** for a product. You will use both the detailed product data from the data entry tool and the top search results retrieved through Azure AI Search from over 5000 indexed products.

### **Objective** ###
- Review the provided product details.
- Analyze the search results that are very similar to the input.
- Precisely determine the product classification focusing on:
    • Class Name  
    • PBH  
    • Analytical Hierarchy (and its corresponding Cd) — note that there are over 3000 options, making accuracy in this field absolutely critical.

### **Classification Criteria** ###
1. **Relevance to Class Name:** How well does the search result match the product category?
2. **Alignment with PBH:** Does the search result fall under the correct product book handler grouping?
3. **Precise Match for Analytical Hierarchy:** Ensure the detailed Analytical Hierarchy and its Cd are correct.

### **Product Information** ###
{product_details}

### **Search Results** ###
"""
    if not search_results:
        prompt += "\n(No search results were found. Proceed using only the product details.)\n"
        logger.warning("No search results found.")
    else:
        for i, result in enumerate(search_results):
            prompt += f"\n#### Match {i+1} ####\n"
            for field in fields:
                prompt += f"- {field.replace('_', ' ').title()}: {result.get(field, '')}\n"
    
    prompt += """

### **Expected Response Format** ###
Respond in exactly the following format:
Class Name: <value>
PBH: <value>
Analytical Hierarchy: <value>
---"""
    logger.info("Prompt to GPT-4o-mini:")
    logger.info(prompt)
    return prompt


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=5, max=60))
def choose_best_prediction(search_results, row):
    """
    Use GPT-4o to select the best prediction based on search results and all relevant fields in the row.
    """
    prompt = build_classification_prompt(row, search_results)

    try:
        response = aoai_client.beta.chat.completions.parse(
            model=openai_model_name,
            messages=[
                {"role": "system", "content": "You are an expert in product categorization."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1000,
            temperature=0,
            response_format=ProductClassification
        )

        prediction = response.choices[0].message.parsed
        return prediction

    except openai.ContentFilterFinishReasonError as e:
        logger.error(f"Content filter error: {e}")
        logger.error(f"Problematic prompt: {prompt}")
        return None

    except Exception as e:
        logger.error(f"An error occurred: {e}")
        logger.error(f"Problematic prompt: {prompt}")
        return None

def predict_classifications(initiated_data, search_client):
    predictions = []
    total_items = len(initiated_data)
    logger.info(f"Total initiated items: {total_items}")  # Print total items for debugging

    for idx, (_, row) in enumerate(initiated_data.iterrows(), start=1):
        # Perform hybrid search
        search_results = hybrid_search(row, search_client)

        # logger.info(f"Search results for item {row['Item_Num']}:")
        # results_exist = False
        # for result in search_results:
        #     logger.info(result)
        #     results_exist = True

        # if not results_exist:
        #     logger.info(f"No search results for item {row['Item_Num']}.")

        # Generate the best prediction (commented out for debugging purposes)
        prediction = choose_best_prediction(search_results, row)
        if prediction:
            predictions.append({
                "Item_Num": row["Item_Num"],
                "Class_Name": prediction.class_name,
                "PBH": prediction.pbh,
                "Analytical_Hierarchy": prediction.analytical_hierarchy
            })

        # Log progress
        if idx % 10 == 0 or idx == total_items:
            logger.info(f"Processed {idx}/{total_items} items")

    return pd.DataFrame(predictions)


def main():
    # File paths
    
    # item_hierarchy_file = "Item Hierarchy 11.14.2024.xlsx"
    # product_sku_items_file = "October Items.xlsx"
    class_pbh_file = "Class PBH.xlsx" # Has the Class, PBH
    item_hierarchy_file = "UPDATED HIERARCHY LIST.xlsx" # Has the Analytical Hierarchy
    product_sku_items_file = "Initiated and Approved Items V2.xlsx" # Has the Approved and Initiated items for a duration of time (i.e. weeks, months)

    # # Step 1: Load the data, get the approved and initiated items
    # class_pbh_df, item_hierarchy_df, approved_items_df, initiated_items_df = load_data(
    #     class_pbh_file, item_hierarchy_file, product_sku_items_file
    # )
    # approved_items_df.to_excel(os.path.join(output_dir, "approved_items.xlsx"), index=False, engine='openpyxl')
    # initiated_items_df.to_excel(os.path.join(output_dir, "initiated_items.xlsx"), index=False, engine='openpyxl')
    
    # Reload the approved and initiated items from the saved files
    approved_items_df = pd.read_excel(os.path.join(output_dir, "approved_items.xlsx"), engine="openpyxl")
    initiated_items_df = pd.read_excel(os.path.join(output_dir, "initiated_items.xlsx"), engine="openpyxl")
    
    
    # Step 2: Build an index of items with embeddings around the Approved items
    # logger.info(f"Number of approved items: {len(approved_items_df)}")
    # create_embedding_index(approved_items_df, search_client)
    # # TODO break up create embedding index function into creating the embeddings, preparing the documents,  creating the index, and uploading the documents.

    # Step 3: Predict the Class, PBH, Analytical_Hierarchy, and save the results
    # initiated_items = pd.read_excel(os.path.join(output_dir, "initiated_items.xlsx"), engine="openpyxl")
    # predictions = predict_classifications(initiated_items, search_client)
    # predictions.to_excel(os.path.join(output_dir, "predicted_results.xlsx"), index=False, engine="openpyxl")

    # Step 4: Evaluate the predictions
    predictions_df = pd.read_excel(os.path.join(output_dir, "predicted_results.xlsx"), engine="openpyxl")
    accuracy = evaluate_predictions(initiated_items_df, predictions_df, approved_items_df)
    logger.info("Accuracy Summary:")
    logger.info(accuracy)

if __name__ == "__main__":
    main()
